from pathlib import Path
import csv

import win32api
import win32print
import win32com.client

import openpyxl


def main() -> None:
    try:
        if win32com.client.GetObject(Class="Excel.Application"):
            print("Close all Excel applications!")
            input("please press the 'ENTER' key to exit.")
            raise RuntimeError("Close all Excel applications!")
    except win32com.client.pywintypes.com_error:
        pass

    p = Path("log.csv")
    # px = p.with_suffix(".xlsx")
    base = Path("base.xlsx")
    px = Path("x.xlsx")

    wb = openpyxl.load_workbook(base)
    ws = WorkSheet(wb["Sheet1"])

    for row in read_csv(p):
        ws.append(row)

    ws.adjust_width()
    ws.setup_page()

    wb.save(px)

    print_with_excel(px)


class WorkSheet:
    def __init__(self, ws) -> None:
        self._ws = ws

    def adjust_width(self) -> None:
        for col in self._ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 1
            self._ws.column_dimensions[column].width = adjusted_width

    def append(self, row) -> None:
        self._ws.append(list(self._convert_row(row)))

    @staticmethod
    def _convert_row(row):
        if ":" in row[0]:
            yield row[0]
            yield row[1]
            yield row[2]
            yield float(row[3])
            yield row[4] if row[4] == "" else float(row[4])
        else:
            for i in row:
                yield i

    def setup_page(self) -> None:
        self._ws.page_setup.fitToWidth = 1
        self._ws.page_setup.fitToHeight = 1
        self._ws.sheet_properties.pageSetUpPr.fitToPage = True


def print_with_excel(p: Path) -> None:
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        ws = excel.Workbooks.Open(p.resolve())
        ws.PrintOut()
    finally:
        ws.Close(SaveChanges=False)
        excel.Quit()


def read_csv(p: Path):
    with open(p, "r", encoding="cp932") as finn:
        for i in csv.reader(finn, quoting=csv.QUOTE_NONNUMERIC):
            yield i


# def set_view(p: Path, px: Path) -> None:
#    print("set_active_sheet")
#    wb = openpyxl.load_workbook(p)
#
#    ws = wb["Sheet1"]
#    for col in ws.columns:
#        max_length = 0
#        column = col[0].column_letter
#        for cell in col:
#            try:
#                if len(str(cell.value)) > max_length:
#                    max_length = len(cell.value)
#            except:
#                pass
#        adjusted_width = (max_length + 1) * 2
#        ws.column_dimensions[column].width = adjusted_width
#
#    ws.page_setup.fitToWidth = 1
#    ws.page_setup.fitToHeight = 1
#    wb.save(px)
# wb.close()

# wb = openpyxl.load_workbook(p)
# ws = wb.worksheets[sheet_number]
# ws.active = sheet_number
# ws.sheet_view.tabSelected = True
# wb.save(p)
# wb.close()


# def print(p: Path) -> None:
#    print("print_excel")
#    win32api.ShellExecute(
#        0, "print", str(p), "/d:" "%s" % win32print.GetDefaultPrinter(), ".", 0
#    )


if __name__ == "__main__":
    main()
